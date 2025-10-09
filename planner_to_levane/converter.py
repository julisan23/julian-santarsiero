"""Conversión de exportaciones de Microsoft Planner a plantillas de Levane."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, MutableMapping, Optional

import pandas as pd
from openpyxl import load_workbook

from .transforms import TRANSFORMS, TransformFunc


@dataclass
class ColumnMapping:
    """Configuración para mapear una columna del Planner al template."""

    template_header: str
    planner_field: Optional[str] = None
    transform: str = "identity"
    options: Mapping[str, Any] | None = None


@dataclass
class ConverterConfig:
    """Configuración completa para ejecutar el convertidor."""

    sheet_name: str
    header_row: int
    mappings: List[ColumnMapping]
    clear_existing: bool = True


def load_configuration(config: Mapping[str, Any]) -> ConverterConfig:
    """Crea un :class:`ConverterConfig` a partir de un diccionario."""

    mappings = [
        ColumnMapping(
            template_header=entry["template_header"],
            planner_field=entry.get("planner_field"),
            transform=entry.get("transform", "identity"),
            options=entry.get("options"),
        )
        for entry in config["mappings"]
    ]

    return ConverterConfig(
        sheet_name=config["sheet_name"],
        header_row=int(config["header_row"]),
        clear_existing=bool(config.get("clear_existing", True)),
        mappings=mappings,
    )


class PlannerToLevaneConverter:
    """Implementa la lógica de conversión entre Planner y Levane."""

    def __init__(self, configuration: ConverterConfig) -> None:
        self.configuration = configuration

    def convert(
        self,
        planner_path: str | Path,
        template_path: str | Path,
        output_path: str | Path,
        *,
        sheet_filters: Optional[Mapping[str, Iterable[Any]]] = None,
    ) -> Path:
        """Realiza la conversión.

        Parameters
        ----------
        planner_path:
            Ruta al archivo de Excel exportado desde Planner.
        template_path:
            Ruta al archivo de plantilla de Levane. No se modifica el archivo
            original; se genera una copia con los datos proyectados.
        output_path:
            Ruta donde se guardará la plantilla con los datos completados.
        sheet_filters:
            Opcionalmente, filtros a aplicar sobre las filas del Planner antes
            de volcarlas en la plantilla.
        """

        planner_path = Path(planner_path)
        template_path = Path(template_path)
        output_path = Path(output_path)

        df = self._load_planner_data(planner_path)
        if sheet_filters:
            df = self._apply_filters(df, sheet_filters)

        workbook = load_workbook(template_path)
        worksheet = workbook[self.configuration.sheet_name]

        header_index = self._build_header_index(worksheet)

        start_row = self.configuration.header_row + 1
        if self.configuration.clear_existing:
            self._clear_existing_rows(worksheet, start_row, header_index.values())

        for offset, (_, row) in enumerate(df.iterrows()):
            excel_row = start_row + offset
            for mapping in self.configuration.mappings:
                column = header_index.get(mapping.template_header)
                if column is None:
                    raise KeyError(
                        f"No se encontró la columna '{mapping.template_header}' en la fila de cabecera"
                    )

                transform = self._resolve_transform(mapping.transform)
                value = transform(
                    row=row,
                    planner_field=mapping.planner_field,
                    options=mapping.options or {},
                )
                worksheet.cell(row=excel_row, column=column, value=value)

        workbook.save(output_path)
        return output_path

    def _load_planner_data(self, planner_path: Path) -> pd.DataFrame:
        df = pd.read_excel(planner_path, dtype=str, keep_default_na=False)
        return df

    def _apply_filters(
        self, df: pd.DataFrame, filters: Mapping[str, Iterable[Any]]
    ) -> pd.DataFrame:
        filtered_df = df.copy()
        for column, allowed_values in filters.items():
            allowed_set = set(allowed_values)
            filtered_df = filtered_df[filtered_df[column].isin(allowed_set)]
        filtered_df.reset_index(drop=True, inplace=True)
        return filtered_df

    def _build_header_index(self, worksheet) -> MutableMapping[str, int]:
        header_index: Dict[str, int] = {}
        for cell in worksheet[self.configuration.header_row]:
            header_value = str(cell.value).strip() if cell.value is not None else None
            if header_value:
                header_index[header_value] = cell.column
        return header_index

    def _clear_existing_rows(
        self, worksheet, start_row: int, columns: Iterable[int]
    ) -> None:
        max_row = worksheet.max_row
        for row in range(start_row, max_row + 1):
            for column in columns:
                worksheet.cell(row=row, column=column, value=None)

    def _resolve_transform(self, name: str) -> TransformFunc:
        try:
            return TRANSFORMS[name]
        except KeyError as exc:
            available = ", ".join(sorted(TRANSFORMS))
            raise KeyError(
                f"Transformación '{name}' no disponible. Transformaciones: {available}"
            ) from exc


